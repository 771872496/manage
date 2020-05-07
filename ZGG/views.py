import json
import os
from datetime import datetime, timezone, timedelta

from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db import connection

from openpyxl import Workbook
from django.utils.http import urlquote
import xlrd

from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect
from io import BytesIO

from ZGG.models import User, Bill, Authority
from django.shortcuts import render_to_response


def page_not_found(request):
    return render_to_response('error/404.html')


def page_error(request):
    return render_to_response('error/404.html')


def login(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        name = name.strip()
        password = request.POST.get('password')
        ret = User.objects.filter(name=name, password=password)
        if ret.exists():
            login_user = ret.first()
            if login_user.state == '禁用':
                error = '该用户已被禁用'
                return render(request, 'login.html', locals())
            else:
                request.session.set_expiry(0)
                request.session['login_user'] = {
                    'user_id': login_user.user_id,
                    'name': login_user.name,
                    'ltt_authority': login_user.ltt.a_name,
                    'zgg_authority': login_user.zgg.a_name,
                    'state': login_user.state
                }
                return redirect('/zgg/bill')
        else:
            error = '账号密码错误,请重新输入'
            return render(request, 'login.html', locals())
    else:
        return render(request, 'login.html', locals())


def admin_logout(request):
    if request.session.get('login_user'):
        del request.session['login_user']
    return redirect('/')


def zgg_user_json(request):
    if request.method == 'GET':
        pageSize = int(request.GET.get('pageSize'))
        pageNumber = int(request.GET.get('pageNumber'))

        rows = []
        total = User.objects.filter(~Q(zgg__a_name='系统管理员')).count()
        user = User.objects.filter(~Q(zgg__a_name='系统管理员'))[(pageNumber - 1) * pageSize:(pageNumber) * pageSize]

        result = {"total": total, "rows": rows}
        for data in user:
            rows.append(
                {'id': data.user_id, 'name': data.name, 'password': data.password, 'state': data.state,
                 'zgg_authority': data.zgg.a_name})

        return HttpResponse(json.dumps(result), content_type="application/json")


def zgg_user(request):
    login_user = request.session['login_user']
    authority = Authority.objects.filter(~Q(a_name='系统管理员'))
    if request.method == 'POST':
        name = request.POST.get('name')
        password = request.POST.get('password')
        ltt = request.POST.get('ltt_id')
        zgg = request.POST.get('zgg_id')
        state = request.POST.get('state')
        find_data = User.objects.filter(name=name)
        if any((find_data, len(password) < 5, not name)):
            error = '该已存在用户'
            return redirect('/zgg/user')
        else:
            User.objects.create(
                name=name,
                password=password,
                state=state,
                ltt_id=ltt,
                zgg_id=zgg,
            )
            error = '创建成功'
        return redirect('/zgg/user')
    return render(request, 'user/zgg_user.html', locals())


def ltt_user_json(request):
    if request.method == 'GET':
        pageSize = int(request.GET.get('pageSize'))
        pageNumber = int(request.GET.get('pageNumber'))

        rows = []
        total = User.objects.filter(~Q(ltt__a_name='系统管理员')).count()
        user = User.objects.filter(~Q(ltt__a_name='系统管理员'))[(pageNumber - 1) * pageSize:(pageNumber) * pageSize]

        result = {"total": total, "rows": rows}
        for data in user:
            rows.append(
                {'id': data.user_id, 'name': data.name, 'password': data.password, 'state': data.state,
                 'ltt_authority': data.ltt.a_name})

        return HttpResponse(json.dumps(result), content_type="application/json")


def ltt_user(request):
    login_user = request.session['login_user']
    authority = Authority.objects.filter(~Q(a_name='系统管理员'))
    if request.method == 'POST':
        name = request.POST.get('name')
        password = request.POST.get('password')
        ltt = request.POST.get('ltt_id')
        zgg = request.POST.get('zgg_id')
        state = request.POST.get('state')
        find_data = User.objects.filter(name=name)
        if any((find_data, len(password) < 5, not name)):
            error = '该已存在用户'
            return redirect('/zgg/user')
        else:
            User.objects.create(
                name=name,
                password=password,
                state=state,
                ltt_id=ltt,
                zgg_id=zgg,
            )
            error = '创建成功'
        return redirect('/ltt/user')
    return render(request, 'user/ltt_user.html', locals())


# 修改用户
def zgg_user_edit(request, user_id):
    login_user = request.session['login_user']
    authority = Authority.objects.filter(~Q(a_name='系统管理员'))
    user = User.objects.filter(pk=user_id).first()
    if request.method == 'POST':
        name = request.POST.get('name')
        password = request.POST.get('password')
        ltt = request.POST.get('ltt_id')
        zgg = request.POST.get('zgg_id')
        state = request.POST.get('state')

        if len(password) < 5:
            error = '密码过短,请重新输入!'
        elif not name:
            error = '用户名不能为空!'
        elif user:
            user.name = name
            user.password = password
            user.ltt_id = ltt
            user.zgg_id = zgg
            user.state = state
            user.save()
            error = '用户信息修改成功!'
        return render(request, 'user/zgg_user_edit.html', locals())
    else:
        return render(request, 'user/zgg_user_edit.html', locals())


def ltt_user_edit(request, user_id):
    login_user = request.session['login_user']
    authority = Authority.objects.filter(~Q(a_name='系统管理员'))
    user = User.objects.filter(pk=user_id).first()
    if request.method == 'POST':
        name = request.POST.get('name')
        password = request.POST.get('password')
        ltt = request.POST.get('ltt_id')
        zgg = request.POST.get('zgg_id')
        state = request.POST.get('state')

        if len(password) < 5:
            error = '密码过短,请重新输入!'
        elif not name:
            error = '用户名不能为空!'
        elif user:
            user.name = name
            user.password = password
            user.ltt_id = ltt
            user.zgg_id = zgg
            user.state = state
            user.save()
            error = '用户信息修改成功!'
        return render(request, 'user/ltt_user_edit.html', locals())
    else:
        return render(request, 'user/ltt_user_edit.html', locals())


def zgg_bill_json(request):
    if request.method == 'GET':
        pageSize = int(request.GET.get('pageSize'))
        pageNumber = int(request.GET.get('pageNumber'))

        zc_number = request.GET.get('search_zc_number').strip()
        zc_site = request.GET.get('search_zc_site').strip()
        oa_id = request.GET.get('search_oa_id').strip()
        state = request.GET.get('search_state')
        rows = []

        if zc_number or zc_site or oa_id or state:
            total = Bill.objects.filter(
                Q(zc_number__contains=zc_number) & Q(zc_site__contains=zc_site) & Q(oa_id__contains=oa_id) & Q(
                    state__contains=state)).order_by('-bill_id').count()
            bill = Bill.objects.filter(
                Q(zc_number__contains=zc_number) & Q(zc_site__contains=zc_site) & Q(oa_id__contains=oa_id) & Q(
                    state__contains=state)).order_by('-bill_id')[(pageNumber - 1) * pageSize:(pageNumber) * pageSize]
        else:
            total = Bill.objects.all().count()
            bill = Bill.objects.order_by('-bill_id')[(pageNumber - 1) * pageSize:(pageNumber) * pageSize]

        result = {"total": total, "rows": rows}
        for data in bill:
            start_time = data.start_time.strftime("%Y-%m-%d %H:%M:%S")
            now_time = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=8)))
            use_time = (now_time - data.start_time).days
            rows.append(
                {'id': data.bill_id, 'zc_type': data.zc_type, 'zc_number': data.zc_number, 'zc_name': data.zc_name,
                 'pro1': data.pro1, 'pro2': data.pro2, 'pro3': data.pro3, 'zc_site': data.zc_site,
                 'oa_id': data.oa_id, 'receive_name': data.receive_name, 'tag': data.tag,
                 'use_time': str(use_time) + '天',
                 'reception_bill': str(data.reception_bill), 'other_bill': str(data.other_bill),
                 'state': data.state, 'execute_user': data.execute_user, 'start_time': str(start_time)})

        return HttpResponse(json.dumps(result), content_type="application/json")


def zgg_bill(request):
    login_user = request.session['login_user']
    if request.method == 'POST':
        zc_type = request.POST.get('zc_type')
        zc_number = request.POST.get('zc_number')
        zc_name = request.POST.get('zc_name')
        pro1 = request.POST.get('pro1')
        pro2 = request.POST.get('pro2')
        pro3 = request.POST.get('pro3')
        zc_site = request.POST.get('zc_site')
        oa_id = request.POST.get('oa_id')
        reception_bill = request.FILES.get('reception_bill')
        now_time = datetime.now(timezone.utc)
        start_time = now_time.astimezone(timezone(timedelta(hours=8)))
        tag = request.POST.get('tag')
        receive_name = request.POST.get('receive_name')
        other_bill = request.FILES.get('other_bill')
        state = request.POST.get('state')
        user_id = request.session['login_user']['user_id']
        execute_user = request.session['login_user']['name']
        find_data = Bill.objects.filter(zc_number=zc_number)

        if find_data.exists():
            error = '该表单已存在'
            return redirect('/zgg/bill', locals())
        else:
            Bill.objects.create(
                zc_type=zc_type,
                zc_number=zc_number,
                zc_name=zc_name,
                pro1=pro1,
                pro2=pro2,
                pro3=pro3,
                zc_site=zc_site,
                oa_id=oa_id,
                reception_bill=reception_bill,
                start_time=start_time,
                tag=tag,
                receive_name=receive_name,
                other_bill=other_bill,
                state=state,
                user_id=user_id,
                execute_user=execute_user
            )
            return redirect('/zgg/bill')
    if request.session['login_user']['zgg_authority'] == '普通用户':
        return render(request, 'zgg/bill_user.html', locals())
    else:
        return render(request, 'zgg/bill.html', locals())


# 资产信息详情
def detail(request, bill_id):
    bill = Bill.objects.filter(pk=bill_id).first()
    login_user = request.session['login_user']
    if request.method == 'GET':
        bill_id = request.POST.get('bill_id')
        zc_type = request.POST.get('zc_type')
        zc_number = request.POST.get('zc_number')
        zc_name = request.POST.get('zc_name')
        pro1 = request.POST.get('pro1')
        pro2 = request.POST.get('pro2')
        pro3 = request.POST.get('pro3')
        zc_site = request.POST.get('zc_site')
        start_time = request.POST.get('start_time')
        oa_id = request.POST.get('oa_id')
        reception_bill = request.POST.get('reception_bill')
        tag = request.POST.get('tag')
        receive_name = request.POST.get('receive_name')
        other_bill = request.POST.get('other_bill')
        state = request.POST.get('state')
        execute_user = request.POST.get('execute_user')
        return render(request, 'zgg/detail.html', locals())
    else:
        return render(request, 'zgg/detail.html', locals())


# 修改资产信息
def bill_edit(request, bill_id):
    bill = Bill.objects.filter(pk=bill_id).first()
    login_user = request.session['login_user']
    if request.method == 'POST':
        zc_type = request.POST.get('zc_type')
        zc_number = request.POST.get('zc_number')
        zc_name = request.POST.get('zc_name')
        pro1 = request.POST.get('pro1')
        pro2 = request.POST.get('pro2')
        pro3 = request.POST.get('pro3')
        zc_site = request.POST.get('zc_site')
        oa_id = request.POST.get('oa_id')
        tag = request.POST.get('tag')
        receive_name = request.POST.get('receive_name')
        state = request.POST.get('state')
        execute_user = request.POST.get('execute_user')
        other_bill: InMemoryUploadedFile = request.FILES.get('other_bill', bill.other_bill)
        reception_bill: InMemoryUploadedFile = request.FILES.get('reception_bill', bill.reception_bill)
        if bill.reception_bill and reception_bill != bill.reception_bill:
            d = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            old_path = os.path.join(d, 'media/%s' % bill.reception_bill.name)
            os.remove(old_path)
        if bill.other_bill and other_bill != bill.other_bill:
            d = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            old_path = os.path.join(d, 'media/%s' % bill.other_bill.name)
            os.remove(old_path)
        if bill:
            bill.zc_type = zc_type
            bill.zc_number = zc_number
            bill.zc_name = zc_name
            bill.pro1 = pro1
            bill.pro2 = pro2
            bill.pro3 = pro3
            bill.zc_site = zc_site
            bill.oa_id = oa_id
            bill.reception_bill = reception_bill
            bill.tag = tag
            bill.receive_name = receive_name
            bill.other_bill = other_bill
            bill.state = state
            bill.execute_bill = execute_user
            bill.save()
            error = '修改成功'
            return redirect('/zgg/bill')
        else:
            error = '修改失败'
    if request.session['login_user']['zgg_authority'] == '普通用户':
        return render(request, 'zgg/bill_edit_admin.html', locals())
    else:
        return render(request, 'zgg/bill_edit.html', locals())


# 批量导入SQL
def import_sql(request):
    login_user = request.session['login_user']
    if request.method == 'POST':
        cursor = connection.cursor()
        try:
            file = request.FILES.get('file', None)
            if not file:
                error = '未选择文件！'
                return render(request, 'zgg/import_sql.html', locals())
            else:
                excel_type = file.name.split('.')[1]

            if excel_type in ['xlsx', 'xls']:
                wb = xlrd.open_workbook(filename=None, file_contents=file.read())
                table = wb.sheets()[0]
                rows = table.nrows
                excel_list = []
                for i in range(1, rows):
                    row = table.row_values(i)
                    now_time = datetime.now(timezone.utc)
                    start_time = now_time.astimezone(timezone(timedelta(hours=8)))
                    user_id = request.session['login_user']['user_id']
                    values = (row[0], row[1], row[2].replace(' ', ''), row[3], row[4],
                              row[5], row[6], row[7], row[8], start_time,
                              row[9], row[10], row[11], row[12], row[13], user_id)
                    zc_number = Bill.objects.filter(zc_number=row[1])

                    if not zc_number:
                        connection.rollback()
                        excel_list.append(values)
                sql = 'INSERT INTO bill(zc_type, zc_number, zc_name, pro1, pro2, ' \
                      'pro3, zc_site, oa_id, reception_bill, start_time, ' \
                      'execute_user, tag, state, receive_name, other_bill, user_id) value ' \
                      '(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
                cursor.executemany(sql, excel_list)
                connection.commit()
                cursor.close()
                error = '数据导入成功！'
                return render(request, 'zgg/import_sql.html', locals())
            else:
                error = '文件选择非Excel文件！'
                return render(request, 'zgg/import_sql.html', locals())
        except Exception:
            return redirect('/zgg/import_sql')
    else:
        return render(request, 'zgg/import_sql.html', locals())


# 导出Excel文件
def export_excel(request):
    wb = Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    row_one = ['资产类别', '固定资产编号', '固定资产名称', '属性1', '属性2', '属性3', '资产地点', '请求ID',
               '接收单', '操作人员', '是否贴标签', '状态', '领用人', '申请单', '录入时间']
    for i in range(1, len(row_one) + 1):  # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    state = request.session['state']['state']
    all_obj = Bill.objects.filter(state=state).all()
    for obj in all_obj:
        max_row = sheet1.max_row + 1  # 获取到工作表的最大行数并加1
        obj_info = [obj.zc_type, obj.zc_number, obj.zc_name, obj.pro1, obj.pro2,
                    obj.pro3, obj.zc_site, obj.oa_id, str(obj.reception_bill),
                    obj.execute_user, obj.tag, obj.state, obj.receive_name, str(obj.other_bill), obj.start_time]

        for i in range(1, len(obj_info) + 1):  # 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=i).value = obj_info[i - 1]
    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = '治具信息%s.xlsx' % ctime  # 给文件名中添加日期时间
    file_name = urlquote(file_name)  # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    # response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
    return response

